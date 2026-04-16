import io
from datetime import datetime, date, timedelta
from flask import Blueprint, request, send_file
from flask_jwt_extended import jwt_required
from backend.app import db
from backend.app.models.g6_hoi_vien import G6HoiVien, G6DangKyGoiTap
from backend.app.models.g6_thanh_toan import G6ThanhToan
from backend.app.utils.g6_phan_hoi import nqt_ok, nqt_loi
from backend.app.utils.g6_xac_thuc import nqt_yeu_cau_dang_nhap
from sqlalchemy import func, extract

nqt_bao_cao_bp = Blueprint('nqt_bao_cao', __name__)


def _nqt_tao_workbook_style(nqt_wb, nqt_tieu_de_sheet: str):
    """Tạo workbook openpyxl với style chuẩn."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    nqt_ws = nqt_wb.active
    nqt_ws.title = nqt_tieu_de_sheet

    nqt_style_header = {
        'font': Font(bold=True, color='FFFFFF', size=11),
        'fill': PatternFill('solid', fgColor='1E3A5F'),
        'alignment': Alignment(horizontal='center', vertical='center', wrap_text=True),
        'border': Border(
            bottom=Side(style='thin', color='CCCCCC'),
            right=Side(style='thin', color='CCCCCC'),
        ),
    }
    return nqt_ws, nqt_style_header


def _nqt_ap_style_header(nqt_ws, nqt_style: dict, nqt_hang: int, nqt_so_cot: int):
    from openpyxl.styles import Font, PatternFill, Alignment, Border
    for nqt_col in range(1, nqt_so_cot + 1):
        nqt_cell = nqt_ws.cell(row=nqt_hang, column=nqt_col)
        nqt_cell.font = nqt_style['font']
        nqt_cell.fill = nqt_style['fill']
        nqt_cell.alignment = nqt_style['alignment']
        nqt_cell.border = nqt_style['border']


# ── 1. Export Excel doanh thu ─────────────────────────────────────────────────

@nqt_bao_cao_bp.route('/api/nqt-bao-cao/doanh-thu/excel', methods=['GET'])
@nqt_yeu_cau_dang_nhap
def nqt_export_doanh_thu_excel():
    """
    Export báo cáo doanh thu theo tháng ra file Excel.
    Query params: nam (int, mặc định năm hiện tại)
    """
    try:
        import openpyxl
    except ImportError:
        return nqt_loi('Thư viện openpyxl chưa được cài đặt.', 500)

    nqt_nam = request.args.get('nam', date.today().year, type=int)

    # Tổng doanh thu theo tháng (từ thanh toán đã xác nhận)
    nqt_ket_qua = (
        db.session.query(
            extract('month', G6ThanhToan.g6_ngay_thanh_toan).label('nqt_thang'),
            func.count(G6ThanhToan.g6_ma_thanh_toan).label('nqt_so_giao_dich'),
            func.sum(G6ThanhToan.g6_so_tien).label('nqt_tong_tien'),
        )
        .filter(
            G6ThanhToan.g6_trang_thai == 'da_thanh_toan',
            extract('year', G6ThanhToan.g6_ngay_thanh_toan) == nqt_nam,
        )
        .group_by(extract('month', G6ThanhToan.g6_ngay_thanh_toan))
        .order_by('nqt_thang')
        .all()
    )

    nqt_wb = openpyxl.Workbook()
    nqt_ws, nqt_style = _nqt_tao_workbook_style(nqt_wb, f'Doanh thu {nqt_nam}')

    # Tiêu đề báo cáo
    nqt_ws.merge_cells('A1:D1')
    nqt_ws['A1'] = f'BÁO CÁO DOANH THU NĂM {nqt_nam}'
    nqt_ws['A1'].font = __import__('openpyxl').styles.Font(bold=True, size=14)
    nqt_ws['A1'].alignment = __import__('openpyxl').styles.Alignment(horizontal='center')

    nqt_ws.merge_cells('A2:D2')
    nqt_ws['A2'] = f'Xuất ngày: {date.today().strftime("%d/%m/%Y")}'
    nqt_ws['A2'].alignment = __import__('openpyxl').styles.Alignment(horizontal='center')

    # Header bảng
    nqt_headers = ['Tháng', 'Số giao dịch', 'Doanh thu (VNĐ)', 'Ghi chú']
    for nqt_i, nqt_h in enumerate(nqt_headers, 1):
        nqt_ws.cell(row=4, column=nqt_i, value=nqt_h)
    _nqt_ap_style_header(nqt_ws, nqt_style, 4, len(nqt_headers))

    nqt_thang_map = {r.nqt_thang: r for r in nqt_ket_qua}
    nqt_tong_nam = 0

    for nqt_thang in range(1, 13):
        nqt_r = nqt_thang_map.get(nqt_thang)
        nqt_so_gd = nqt_r.nqt_so_giao_dich if nqt_r else 0
        nqt_tien = float(nqt_r.nqt_tong_tien) if nqt_r else 0
        nqt_tong_nam += nqt_tien
        nqt_row = nqt_thang + 4
        nqt_ws.cell(row=nqt_row, column=1, value=f'Tháng {nqt_thang:02d}')
        nqt_ws.cell(row=nqt_row, column=2, value=nqt_so_gd)
        nqt_ws.cell(row=nqt_row, column=3, value=nqt_tien)
        nqt_ws.cell(row=nqt_row, column=3).number_format = '#,##0'

    # Dòng tổng
    nqt_row_tong = 17
    nqt_ws.cell(row=nqt_row_tong, column=1, value='TỔNG NĂM')
    nqt_ws.cell(row=nqt_row_tong, column=1).font = __import__('openpyxl').styles.Font(bold=True)
    nqt_ws.cell(row=nqt_row_tong, column=3, value=nqt_tong_nam)
    nqt_ws.cell(row=nqt_row_tong, column=3).font = __import__('openpyxl').styles.Font(bold=True)
    nqt_ws.cell(row=nqt_row_tong, column=3).number_format = '#,##0'

    nqt_ws.column_dimensions['A'].width = 15
    nqt_ws.column_dimensions['B'].width = 16
    nqt_ws.column_dimensions['C'].width = 22
    nqt_ws.column_dimensions['D'].width = 20

    nqt_buf = io.BytesIO()
    nqt_wb.save(nqt_buf)
    nqt_buf.seek(0)

    return send_file(
        nqt_buf,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'bao_cao_doanh_thu_{nqt_nam}.xlsx',
    )


# ── 2. Export Excel danh sách hội viên ───────────────────────────────────────

@nqt_bao_cao_bp.route('/api/nqt-bao-cao/hoi-vien/excel', methods=['GET'])
@nqt_yeu_cau_dang_nhap
def nqt_export_hoi_vien_excel():
    """Export danh sách toàn bộ hội viên đang hoạt động ra Excel."""
    try:
        import openpyxl
    except ImportError:
        return nqt_loi('Thư viện openpyxl chưa được cài đặt.', 500)

    nqt_danh_sach = (
        G6HoiVien.query
        .filter(G6HoiVien.g6_la_hoat_dong == True)
        .order_by(G6HoiVien.g6_ngay_dang_ky.desc())
        .all()
    )

    nqt_wb = openpyxl.Workbook()
    nqt_ws, nqt_style = _nqt_tao_workbook_style(nqt_wb, 'Hội viên')

    nqt_ws.merge_cells('A1:G1')
    nqt_ws['A1'] = 'DANH SÁCH HỘI VIÊN'
    nqt_ws['A1'].font = __import__('openpyxl').styles.Font(bold=True, size=14)
    nqt_ws['A1'].alignment = __import__('openpyxl').styles.Alignment(horizontal='center')

    nqt_ws.merge_cells('A2:G2')
    nqt_ws['A2'] = f'Xuất ngày: {date.today().strftime("%d/%m/%Y")} — Tổng: {len(nqt_danh_sach)} hội viên'
    nqt_ws['A2'].alignment = __import__('openpyxl').styles.Alignment(horizontal='center')

    nqt_headers = ['STT', 'Mã HV', 'Họ tên', 'Điện thoại', 'Email', 'Ngày đăng ký', 'Trạng thái']
    for nqt_i, nqt_h in enumerate(nqt_headers, 1):
        nqt_ws.cell(row=4, column=nqt_i, value=nqt_h)
    _nqt_ap_style_header(nqt_ws, nqt_style, 4, len(nqt_headers))

    for nqt_i, nqt_hv in enumerate(nqt_danh_sach, 1):
        nqt_row = nqt_i + 4
        nqt_ws.cell(row=nqt_row, column=1, value=nqt_i)
        nqt_ws.cell(row=nqt_row, column=2, value=nqt_hv.g6_ma_hoi_vien)
        nqt_ws.cell(row=nqt_row, column=3, value=nqt_hv.g6_ho_ten)
        nqt_ws.cell(row=nqt_row, column=4, value=nqt_hv.g6_so_dien_thoai)
        nqt_ws.cell(row=nqt_row, column=5, value=nqt_hv.g6_email or '')
        nqt_ws.cell(row=nqt_row, column=6,
                    value=str(nqt_hv.g6_ngay_dang_ky) if nqt_hv.g6_ngay_dang_ky else '')
        nqt_ws.cell(row=nqt_row, column=7, value='Hoạt động' if nqt_hv.g6_la_hoat_dong else 'Ngừng')

    for nqt_col, nqt_w in zip('ABCDEFG', [6, 10, 25, 15, 25, 14, 12]):
        nqt_ws.column_dimensions[nqt_col].width = nqt_w

    nqt_buf = io.BytesIO()
    nqt_wb.save(nqt_buf)
    nqt_buf.seek(0)

    return send_file(
        nqt_buf,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'danh_sach_hoi_vien_{date.today().strftime("%Y%m%d")}.xlsx',
    )


# ── 3. Export Excel gói tập sắp hết hạn ─────────────────────────────────────

@nqt_bao_cao_bp.route('/api/nqt-bao-cao/het-han/excel', methods=['GET'])
@nqt_yeu_cau_dang_nhap
def nqt_export_het_han_excel():
    """
    Export danh sách gói tập sắp hết hạn trong N ngày tới ra Excel.
    Query param: ngay (int, mặc định 30)
    """
    try:
        import openpyxl
    except ImportError:
        return nqt_loi('Thư viện openpyxl chưa được cài đặt.', 500)

    nqt_so_ngay = request.args.get('ngay', 30, type=int)
    nqt_ngay_cuoi = date.today() + timedelta(days=nqt_so_ngay)

    nqt_danh_sach = (
        db.session.query(G6DangKyGoiTap, G6HoiVien)
        .join(G6HoiVien, G6DangKyGoiTap.g6_ma_hoi_vien == G6HoiVien.g6_ma_hoi_vien)
        .filter(
            G6DangKyGoiTap.g6_trang_thai == 'dang_hoat_dong',
            G6DangKyGoiTap.g6_ngay_het_han >= date.today(),
            G6DangKyGoiTap.g6_ngay_het_han <= nqt_ngay_cuoi,
        )
        .order_by(G6DangKyGoiTap.g6_ngay_het_han)
        .all()
    )

    nqt_wb = openpyxl.Workbook()
    nqt_ws, nqt_style = _nqt_tao_workbook_style(nqt_wb, f'Hết hạn {nqt_so_ngay} ngày')

    nqt_ws.merge_cells('A1:F1')
    nqt_ws['A1'] = f'DANH SÁCH GÓI TẬP HẾT HẠN TRONG {nqt_so_ngay} NGÀY TỚI'
    nqt_ws['A1'].font = __import__('openpyxl').styles.Font(bold=True, size=14)
    nqt_ws['A1'].alignment = __import__('openpyxl').styles.Alignment(horizontal='center')

    nqt_ws.merge_cells('A2:F2')
    nqt_ws['A2'] = (
        f'Từ ngày {date.today().strftime("%d/%m/%Y")} '
        f'đến {nqt_ngay_cuoi.strftime("%d/%m/%Y")} — Tổng: {len(nqt_danh_sach)} gói'
    )
    nqt_ws['A2'].alignment = __import__('openpyxl').styles.Alignment(horizontal='center')

    nqt_headers = ['STT', 'Họ tên', 'Điện thoại', 'Tên gói', 'Ngày hết hạn', 'Còn (ngày)']
    for nqt_i, nqt_h in enumerate(nqt_headers, 1):
        nqt_ws.cell(row=4, column=nqt_i, value=nqt_h)
    _nqt_ap_style_header(nqt_ws, nqt_style, 4, len(nqt_headers))

    for nqt_i, (nqt_dk, nqt_hv) in enumerate(nqt_danh_sach, 1):
        nqt_con_lai = (nqt_dk.g6_ngay_het_han - date.today()).days
        nqt_ten_goi = nqt_dk.g6_goi_tap.g6_ten_goi if nqt_dk.g6_goi_tap else f'Gói #{nqt_dk.g6_ma_goi_tap}'
        nqt_row = nqt_i + 4
        nqt_ws.cell(row=nqt_row, column=1, value=nqt_i)
        nqt_ws.cell(row=nqt_row, column=2, value=nqt_hv.g6_ho_ten)
        nqt_ws.cell(row=nqt_row, column=3, value=nqt_hv.g6_so_dien_thoai)
        nqt_ws.cell(row=nqt_row, column=4, value=nqt_ten_goi)
        nqt_ws.cell(row=nqt_row, column=5, value=str(nqt_dk.g6_ngay_het_han))
        nqt_ws.cell(row=nqt_row, column=6, value=nqt_con_lai)

        # Tô màu đỏ nếu còn ≤ 7 ngày
        if nqt_con_lai <= 7:
            nqt_ws.cell(row=nqt_row, column=6).font = (
                __import__('openpyxl').styles.Font(color='CC0000', bold=True)
            )

    for nqt_col, nqt_w in zip('ABCDEF', [6, 25, 15, 22, 14, 12]):
        nqt_ws.column_dimensions[nqt_col].width = nqt_w

    nqt_buf = io.BytesIO()
    nqt_wb.save(nqt_buf)
    nqt_buf.seek(0)

    return send_file(
        nqt_buf,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'goi_tap_het_han_{date.today().strftime("%Y%m%d")}.xlsx',
    )


# ── 4. Tổng hợp số liệu JSON ─────────────────────────────────────────────────

@nqt_bao_cao_bp.route('/api/nqt-bao-cao/tong-hop', methods=['GET'])
@nqt_yeu_cau_dang_nhap
def nqt_tong_hop_bao_cao():
    """Trả về số liệu tổng hợp dạng JSON cho dashboard báo cáo."""
    nqt_hom_nay = date.today()
    nqt_dau_thang = nqt_hom_nay.replace(day=1)

    nqt_tong_hoi_vien = G6HoiVien.query.filter_by(g6_la_hoat_dong=True).count()

    nqt_moi_thang = G6HoiVien.query.filter(
        G6HoiVien.g6_ngay_dang_ky >= nqt_dau_thang,
        G6HoiVien.g6_la_hoat_dong == True,
    ).count()

    nqt_goi_sap_het = G6DangKyGoiTap.query.filter(
        G6DangKyGoiTap.g6_trang_thai == 'dang_hoat_dong',
        G6DangKyGoiTap.g6_ngay_het_han >= nqt_hom_nay,
        G6DangKyGoiTap.g6_ngay_het_han <= nqt_hom_nay + timedelta(days=7),
    ).count()

    nqt_doanh_thu_thang = db.session.query(
        func.sum(G6ThanhToan.g6_so_tien)
    ).filter(
        G6ThanhToan.g6_trang_thai == 'da_thanh_toan',
        G6ThanhToan.g6_ngay_thanh_toan >= nqt_dau_thang,
    ).scalar() or 0

    return nqt_ok({
        'nqt_tong_hoi_vien': nqt_tong_hoi_vien,
        'nqt_hoi_vien_moi_thang': nqt_moi_thang,
        'nqt_goi_sap_het_han_7_ngay': nqt_goi_sap_het,
        'nqt_doanh_thu_thang_nay': float(nqt_doanh_thu_thang),
        'nqt_thang': nqt_hom_nay.month,
        'nqt_nam': nqt_hom_nay.year,
    })
